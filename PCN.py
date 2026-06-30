import os, time, re, datetime, threading, sys, json, base64
import pandas as pd
import tkinter as tk
from tkinter import messagebox, scrolledtext
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
# 使用正確的 LaTeX 格式範例：若需計算進度分比可參考 $P = \frac{n}{N} \times 100\%$
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import SessionNotCreatedException, TimeoutException, WebDriverException
from bs4 import BeautifulSoup
from openpyxl import load_workbook

# --- 路徑設定 ---
if getattr(sys, 'frozen', False):
    BASE_DIR = os.path.dirname(sys.executable)
else:
    BASE_DIR = os.path.dirname(os.path.abspath(__file__))

DATA_DIR = os.path.join(BASE_DIR, "DATA")
os.makedirs(DATA_DIR, exist_ok=True)

PATH_CHROMEDRIVER = os.path.join(DATA_DIR, "chromedriver.exe")
PATH_EXCEL = os.path.join(DATA_DIR, "公版格式.xlsx")
PATH_WAIT_TXT = os.path.join(DATA_DIR, "waitnotdopcn.txt")
PATH_CONFIG = os.path.join(DATA_DIR, "config.json")

class PCNApp:
    def __init__(self, root):
        self.root = root
        self.root.title("PCN 自動化系統 v4.5 (全量記錄版)")
        self.root.geometry("500x680")

        tk.Label(root, text="PCN 系統登入", font=("Arial", 16, "bold")).pack(pady=10)

        tk.Label(root, text="使用者帳號:").pack()
        self.entry_user = tk.Entry(root, width=30)
        self.entry_user.pack(pady=5)

        tk.Label(root, text="使用者密碼:").pack()
        self.entry_pass = tk.Entry(root, width=30, show="*")
        self.entry_pass.pack(pady=5)

        self.remember_var = tk.BooleanVar()
        self.check_remember = tk.Checkbutton(root, text="記住帳號密碼", variable=self.remember_var)
        self.check_remember.pack(pady=2)

        tk.Label(root, text="手動輸入單號 (選填):", fg="#1976D2", font=("Arial", 10, "bold")).pack(pady=(10,0))
        self.entry_manual_pcn = tk.Entry(root, width=30)
        self.entry_manual_pcn.pack(pady=2)
        tk.Label(root, text="(若留空則自動處理所有 WIP 項目)", font=("Arial", 8), fg="gray").pack()

        self.btn_run = tk.Button(root, text="開始執行", command=self.start_thread, bg="#4CAF50", fg="white", width=20, height=2)
        self.btn_run.pack(pady=15)

        tk.Label(root, text="執行狀態:").pack(anchor="w", padx=20)
        self.log_area = scrolledtext.ScrolledText(root, width=60, height=15, state='disabled')
        self.log_area.pack(padx=20, pady=5)
        self.log_area.tag_config("success", foreground="#2E7D32")
        self.log_area.tag_config("error", foreground="#C62828")
        self.log_area.tag_config("warning", foreground="#EF6C00")

        self.load_user_config()

    def load_user_config(self):
        if os.path.exists(PATH_CONFIG):
            try:
                with open(PATH_CONFIG, 'r') as f:
                    data = json.load(f)
                    self.entry_user.insert(0, data.get("user", ""))
                    pwd_encoded = data.get("pass", "")
                    if pwd_encoded:
                        pwd_decoded = base64.b64decode(pwd_encoded).decode('utf-8')
                        self.entry_pass.insert(0, pwd_decoded)
                        self.remember_var.set(True)
            except: pass

    def save_user_config(self, user, pwd):
        if self.remember_var.get():
            pwd_encoded = base64.b64encode(pwd.encode('utf-8')).decode('utf-8')
            config = {"user": user, "pass": pwd_encoded}
            with open(PATH_CONFIG, 'w') as f:
                json.dump(config, f)
        else:
            if os.path.exists(PATH_CONFIG): os.remove(PATH_CONFIG)

    def log(self, message):
        self.log_area.config(state='normal')
        tag = None
        if "✅" in message: tag = "success"
        elif "❌" in message: tag = "error"
        elif "⚠" in message: tag = "warning"

        timestamp = datetime.datetime.now().strftime('%H:%M:%S')
        self.log_area.insert(tk.END, f"[{timestamp}] {message}\n", tag)
        self.log_area.see(tk.END)
        self.log_area.config(state='disabled')
        self.root.update()

    def start_thread(self):
        if not self.entry_user.get() or not self.entry_pass.get():
            messagebox.showwarning("提示", "請輸入帳號與密碼")
            return
        self.btn_run.config(state='disabled', text="執行中...", bg="#9E9E9E")
        threading.Thread(target=self.run_automation, daemon=True).start()

    def run_automation(self):
        user_id = self.entry_user.get().strip()
        user_pass = self.entry_pass.get()
        manual_pcn = self.entry_manual_pcn.get().strip()

        self.save_user_config(user_id, user_pass)

        chrome_options = Options()
        chrome_options.add_argument("--headless")
        chrome_options.add_experimental_option('excludeSwitches', ['enable-logging'])

        driver = None
        try:
            driver = webdriver.Chrome(service=Service(PATH_CHROMEDRIVER), options=chrome_options)
            wait = WebDriverWait(driver, 15)
            mytime = datetime.datetime.now()

            # --- 1. 登入伺服器 ---
            target_ips = ['10.55.7.91', '10.55.7.92']
            current_ip = ""
            for ip in target_ips:
                try:
                    driver.get(f"http://{ip}/PCN/WorklistAction.do?method=dispInitPage")
                    current_ip = ip
                    self.log(f"✅ 已連線至伺服器 {ip}")
                    break
                except WebDriverException:
                    self.log(f"⚠ {ip} 無回應，嘗試下一個...")

            if not current_ip:
                self.log("❌ 錯誤：所有伺服器 IP 皆無法連線。")
                return

            wait.until(EC.presence_of_element_located((By.ID, "tbUserId"))).send_keys(user_id)
            driver.find_element(By.ID, "tbPassword").send_keys(user_pass)
            driver.find_element(By.ID, "btnLogin").click()

            # 檢查是否彈出登入錯誤
            try:
                WebDriverWait(driver, 3).until(EC.alert_is_present())
                alert = driver.switch_to.alert
                self.log(f"❌ 登入失敗：{alert.text}")
                alert.accept()
                return
            except TimeoutException:
                pass

            # --- 2. 獲取單號清單 ---
            wait.until(EC.presence_of_element_located((By.XPATH, "//*")))
            soup = BeautifulSoup(driver.page_source, 'html.parser')
            all_links = soup.find_all(href=re.compile("TW$"))

            process_list = []
            if manual_pcn:
                process_list = [manual_pcn]
                self.log(f"⌨ 手動模式：僅處理 {manual_pcn}")
            else:
                notyet = []
                if os.path.exists(PATH_WAIT_TXT):
                    with open(PATH_WAIT_TXT, 'r', encoding='utf-8') as f:
                        notyet = [line.strip().replace('PCN-', '') for line in f.readlines() if line.strip()]

                # 自動模式：抓取所有未在排除清單中的單號
                for l in all_links[::-1]:
                    pcn_id = l.text.strip()
                    if pcn_id[-10:] not in notyet:
                        process_list.append(pcn_id)

                self.log(f"📋 自動模式：WIP 總筆數 {len(all_links)}，待處理 {len(process_list)} 筆")

            if not process_list:
                self.log("⚠ 狀態：目前沒有需要處理的案件。")
                return

            # --- 3. 逐筆處理並寫入 Excel ---
            for idx, target_pcn in enumerate(process_list, 1):
                try:
                    self.log(f"⏳ ({idx}/{len(process_list)}) 正在抓取：{target_pcn}")
                    driver.get(f'http://{current_ip}/PCN/IssueApplyAction.do?method=dispInitPagesSign&PCN_ID={target_pcn}')

                    # 顯式等待頁面加載
                    wait.until(EC.presence_of_element_located((By.XPATH, "//*")))

                    # 抓取詳細資料
                    status = driver.find_element(By.XPATH, "//strong/font").text.strip()
                    fab = driver.find_element(By.ID, "mainFabSpan").text.strip()
                    subject = driver.find_element(By.XPATH, "//*/following::pre").text.strip()
                    owner = driver.find_element(By.XPATH, "//td[@class='tdcontent002']/following-sibling::td").text.strip()
                    creatdate = driver.find_element(By.XPATH, '//*[@id="Table_01"]/tbody/tr/td/form/div/div/div/div/table[3]/tbody/tr/td/table/tbody/tr[1]/td[6]').text.strip()
                    arrivedate = (lambda xpath: driver.find_element(By.XPATH, xpath).text.strip() if driver.find_elements(By.XPATH, xpath) else "手動輸入單號")('//*[@id="Table_01"]/tbody/tr/td/form/div/div/div/table[3]/tbody/tr[3]/td[3]')
                    # 處理 Excel 寫入
                    if os.path.exists(PATH_EXCEL):
                        try:
                            wb = load_workbook(PATH_EXCEL)
                            sheet_name = "7.簽核文件"
                            if sheet_name not in wb.sheetnames:
                                self.log(f"❌ 錯誤：Excel 中找不到分頁「{sheet_name}」")
                                break

                            ws = wb[sheet_name]

                            # 檢查是否重複（掃描最後 20 列）
                            is_duplicate = False
                            for row in range(max(1, ws.max_row - 20), ws.max_row + 1):
                                if str(ws.cell(row=row, column=6).value) == str(target_pcn):
                                    is_duplicate = True
                                    break

                            if not is_duplicate:
                                is_special = ('Emergency' in status) or manual_pcn
                                new_row = [
                                    mytime.isocalendar()[1], # 週次
                                    '',
                                    '*' if is_special else '',
                                    f"開單{creatdate}到單{arrivedate}紀錄{mytime.strftime('%Y-%m-%d %H:%M')}",
                                    fab,
                                    target_pcn,
                                    f'{owner} PCN的WIP:{len(all_links)}張',
                                    subject,
                                    status,
                                    owner if is_special else ''
                                ]
                                ws.append(new_row)
                                wb.save(PATH_EXCEL)
                                self.log(f"✅ {target_pcn} 記錄成功")
                            else:
                                self.log(f"⏭ {target_pcn} 已在 Excel 中，跳過。")
                        except PermissionError:
                            self.log(f"❌ 錯誤：Excel 檔案被開啟中，無法存檔 {target_pcn}。")
                    else:
                        self.log(f"⚠ 找不到 Excel 檔案：{PATH_EXCEL}")
                        break

                except Exception as e:
                    self.log(f"❌ 處理 {target_pcn} 時發生錯誤：{e}")
                    continue # 發生錯誤則跳下一筆


        except Exception as e:
            self.log(f"❌ 執行中斷：{e}")
        finally:
            if driver:
                driver.quit()
            self.log("🏁 全部作業結束。")
            self.btn_run.config(state='normal', text="開始執行", bg="#4CAF50")
            messagebox.showinfo("完成", f"已完成處理 {len(process_list)} 筆單據！")

if __name__ == "__main__":
    root = tk.Tk()
    app = PCNApp(root)
    root.mainloop()
