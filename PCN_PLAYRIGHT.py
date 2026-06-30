import asyncio
from playwright.async_api import async_playwright
from bs4 import BeautifulSoup
from openpyxl import load_workbook
import datetime, os, re

# --- 原有的路徑設定 ---
DATA_DIR = "DATA"
PATH_EXCEL = os.path.join(DATA_DIR, "公版格式.xlsx")
SHEET_NAME = "7.簽核文件"

async def run_pcn_automation(user_id, user_pass, manual_pcn=None):
    async with async_playwright() as p:
        # 啟動瀏覽器 (headless=True 為無後台模式)
        browser = await p.chromium.launch(headless=True)
        context = await browser.new_context()
        page = await context.new_page()

        try:
            # 1. 登入 (以 10.55.7.91 為例)
            print(f"正在連線至伺服器...")
            await page.goto("http://10.55.7.91")

            await page.fill("#tbUserId", user_id)
            await page.fill("#tbPassword", user_pass)
            await page.click("#btnLogin")

            # 等待跳轉完成
            await page.wait_for_load_state("networkidle")

            # 2. 獲取單據清單
            content = await page.content()
            soup = BeautifulSoup(content, 'html.parser')
            all_links = soup.find_all(href=re.compile("TW$"))

            process_list = [manual_pcn] if manual_pcn else []
            print(f"待處理項目：{len(process_list)} 筆")

            # 3. 逐筆處理
            for idx, target_pcn in enumerate(process_list, 1):
                print(f"({idx}/{len(process_list)}) 處理中：{target_pcn}")

                detail_url = f"http://10.55.7{target_pcn}"
                await page.goto(detail_url)
                await page.wait_for_selector("//strong/font") # 等待關鍵元素出現

                # 抓取資料 (使用 Playwright 的定位器)
                status = await page.locator("//strong/font").inner_text()
                fab = await page.locator("#mainFabSpan").inner_text()
                subject = await page.locator("//*/following::pre").first.inner_text()
                owner = await page.locator("//td[@class='tdcontent002']/following-sibling::td").first.inner_text()

                # --- 寫入 Excel 邏輯 (與你原本相同) ---
                if os.path.exists(PATH_EXCEL):
                    wb = load_workbook(PATH_EXCEL)
                    ws = wb[SHEET_NAME]

                    # 簡單重複檢查
                    is_duplicate = False
                    for row in range(max(1, ws.max_row - 20), ws.max_row + 1):
                        if str(ws.cell(row=row, column=6).value) == str(target_pcn):
                            is_duplicate = True; break

                    if not is_duplicate:
                        mytime = datetime.datetime.now()
                        new_row = [
                            mytime.isocalendar()[1], # 週次
                            '',
                            '*' if 'Emergency' in status else '',
                            f"開單紀全 {mytime.strftime('%Y-%m-%d %H:%M')}",
                            fab,
                            target_pcn,
                            f"{owner} PCN目WIP",
                            subject,
                            status,
                            owner if 'Emergency' in status else ''
                        ]
                        ws.append(new_row)
                        wb.save(PATH_EXCEL)
                        print(f"✅ {target_pcn} 紀錄成功")

        except Exception as e:
            print(f"❌ 發生錯誤: {e}")
        finally:
            await browser.close()

# 執行
if __name__ == "__main__":
    asyncio.run(run_pcn_automation("yongjie.chou", "@S122765079x"))
