import os, time, re, datetime, sys, json, base64
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException, WebDriverException
from bs4 import BeautifulSoup
from openpyxl import load_workbook

# --- 路徑設定 ---
if getattr(sys, 'frozen', False):
    BASE_DIR = os.path.dirname(sys.executable)
else:
    BASE_DIR = os.path.dirname(os.path.abspath(__file__))

DATA_DIR = os.path.join(BASE_DIR, "DATA")
os.makedirs(DATA_DIR, exist_ok=True)
PATH_CHROMEDRIVER = os.path.join(DATA_DIR, "chromedriver.exe")
PATH_EXCEL = os.path.join(DATA_DIR, "公版格式.xlsx")
PATH_WAIT_TXT = os.path.join(DATA_DIR, "waitnotdopcn.txt")
PATH_CONFIG = os.path.join(DATA_DIR, "config.json")

# 🟢 如果你想直接把帳密寫死在程式碼裡，請修改這裡：
DEFAULT_USER = "" # 例如 "A12345"
DEFAULT_PASS = "" # 例如 "Password123"

class PCNAuto:
    def log(self, message):
        timestamp = datetime.datetime.now().strftime('%H:%M:%S')
        print(f"[{timestamp}] {message}")

    def get_credentials(self):
        # 1. 檢查是否有寫死的帳密
        if DEFAULT_USER and DEFAULT_PASS:
            return DEFAULT_USER, DEFAULT_PASS

        # 2. 檢查 config.json
        if os.path.exists(PATH_CONFIG):
            try:
                with open(PATH_CONFIG, 'r') as f:
                    data = json.load(f)
                    user = data.get("user", "")
                    pwd_encoded = data.get("pass", "")
                    pwd = base64.b64decode(pwd_encoded).decode('utf-8')
                    if user and pwd: return user, pwd
            except: pass

        # 3. 若都沒有，才要求輸入一次
        print("首次執行，請設定帳密（之後將自動讀取）：")
        user = input("帳號: ").strip()
        pwd = input("密碼: ").strip()
        pwd_encoded = base64.b64encode(pwd.encode('utf-8')).decode('utf-8')
        with open(PATH_CONFIG, 'w') as f:
            json.dump({"user": user, "pass": pwd_encoded}, f)
        return user, pwd

    def run(self):
        user_id, user_pass = self.get_credentials()
        self.log(f"🚀 啟動自動化系統 (帳號: {user_id})")

        chrome_options = Options()
        chrome_options.add_argument("--headless") # 不顯示瀏覽器視窗
        chrome_options.add_experimental_option('excludeSwitches', ['enable-logging'])

        driver = None
        try:
            driver = webdriver.Chrome(service=Service(PATH_CHROMEDRIVER), options=chrome_options)
            wait = WebDriverWait(driver, 15)
            mytime = datetime.datetime.now()

            # --- 1. 登入 ---
            target_ips = ['10.55.7.91', '10.55.7.92']
            current_ip = ""
            for ip in target_ips:
                try:
                    driver.get(f"http://{ip}/PCN/WorklistAction.do?method=dispInitPage")
                    current_ip = ip
                    self.log(f" ✅ 已連線 {ip}")
                    break
                except WebDriverException:
                    self.log(f" ⚠ {ip} 無回應...")

            if not current_ip:
                self.log(" ❌ 錯誤：無法連線至任何伺服器。")
                return

            wait.until(EC.presence_of_element_located((By.ID, "tbUserId"))).send_keys(user_id)
            driver.find_element(By.ID, "tbPassword").send_keys(user_pass)
            driver.find_element(By.ID, "btnLogin").click()

            # 檢查登入錯誤 (Alert)
            try:
                WebDriverWait(driver, 3).until(EC.alert_is_present())
                alert = driver.switch_to.alert
                self.log(f" ❌ 登入失敗：{alert.text}")
                alert.accept()
                return
            except: pass

            # --- 2. 獲取單號 ---
            wait.until(EC.presence_of_element_located((By.XPATH, "//*")))
            soup = BeautifulSoup(driver.page_source, 'html.parser')
            all_links = soup.find_all(href=re.compile("TW$"))

            notyet = []
            if os.path.exists(PATH_WAIT_TXT):
                with open(PATH_WAIT_TXT, 'r', encoding='utf-8') as f:
                    notyet = [line.strip().replace('PCN-', '') for line in f.readlines() if line.strip()]

            process_list = [l.text.strip() for l in all_links[::-1] if l.text.strip()[-10:] not in notyet]
            self.log(f" 📊 待處理單號：{len(process_list)} 筆")

            if not process_list:
                return

            # --- 3. 處理 Excel ---
            for idx, target_pcn in enumerate(process_list, 1):
                try:
                    self.log(f" ({idx}/{len(process_list)}) ⏳ 處理中：{target_pcn}")
                    driver.get(f'http://{current_ip}/PCN/IssueApplyAction.do?method=dispInitPagesSign&PCN_ID={target_pcn}')
                    wait.until(EC.presence_of_element_located((By.XPATH, "//*")))

                    # 抓取資料
                    status = driver.find_element(By.XPATH, "//strong/font").text.strip()
                    fab = driver.find_element(By.ID, "mainFabSpan").text.strip()
                    subject = driver.find_element(By.XPATH, "//*/following::pre").text.strip()
                    owner = driver.find_element(By.XPATH, "//td[@class='tdcontent002']/following-sibling::td").text.strip()
                    creatdate = driver.find_element(By.XPATH, '//*[@id="Table_01"]/tbody/tr/td/form/div/div/div/div/table/tbody/tr/td/table/tbody/tr/td').text.strip()

                    if os.path.exists(PATH_EXCEL):
                        wb = load_workbook(PATH_EXCEL)
                        ws = wb["7.簽核文件"]

                        # 簡單檢查重複 (最後20列)
                        is_duplicate = any(str(ws.cell(row=r, column=6).value) == str(target_pcn) for r in range(max(1, ws.max_row-20), ws.max_row+1))

                        if not is_duplicate:
                            new_row = [
                                mytime.isocalendar()[1], # 週次
                                '',
                                '*' if 'Emergency' in status else '',
                                f"紀錄時間 {mytime.strftime('%Y-%m-%d %H:%M')}",
                                fab,
                                target_pcn,
                                f'{owner} PCN單WIP:{len(all_links)}張',
                                subject,
                                status,
                                owner if 'Emergency' in status else ''
                            ]
                            ws.append(new_row)
                            wb.save(PATH_EXCEL)
                            self.log(f"    ✅ 寫入成功")
                        else:
                            self.log(f"    ⏭ 跳過重複")
                except Exception as e:
                    self.log(f"    ❌ 錯誤：{e}")

        except Exception as e:
            self.log(f" ❌ 中斷：{e}")
        finally:
            if driver: driver.quit()
            self.log(" 🏁 作業完成。")
            time.sleep(3) # 留 3 秒看結果後自動關閉

if __name__ == "__main__":
    app = PCNAuto()
    app.run()
