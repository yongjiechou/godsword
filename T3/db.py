import os
import pandas as pd
import datetime
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
##pd.set_option('display.max_rows', None)
pd.set_option('display.max_columns', None)
pd.set_option('display.width', None) ##視情況調適欄位對應好看
pd.set_option('display.max_colwidth', None)

os.add_dll_directory(r'T:\Software\WebSphere MQ 6.0_Client\Licenses\Windows\Portable Python-3.9.9 x64\App\Python\Lib\site-packages\clidriver\bin')

import ibm_db
import ibm_db_dbi
# Careful with the punctuation here - we have 3 arguments.
# The first is a big string with semicolons in it.
# (Strings separated by only whitespace, newlines included,
#  are automatically joined together, in case you didn't know.)
# The last two are emptry strings.

def reset_col(filename):
    wb=load_workbook(filename)
    for sheet in wb.sheetnames:
        ws=wb[sheet]
        df=pd.read_excel(filename,sheet).fillna('-')
        df.loc[len(df)]=list(df.columns)#把標題行附加到最後一行
        for col in df.columns:
            index=list(df.columns).index(col)					#列序號
            letter=get_column_letter(index+1)					#列字母
            collen=df[col].apply(lambda x:len(str(x).encode())).max()	#獲取這一列長度的最大值 當然也可以用min獲取最小值 mean獲取平均值
            ws.column_dimensions[letter].width=collen*1.5
            ws.freeze_panes='B3'
        wb.save(filename)
mydic={
'T7':'ARRAY7',
'F7':'CF7',
'L7':'LCD7',
'TP':'T3',

}

cnxntft = ibm_db.connect("DATABASE=TPABRM;HOSTNAME=10.32.201.31;PORT=54150;PROTOCOL=TCPIP;UID=t3apbrm;PWD=t3apbrm;", "", "")


mytime=datetime.datetime.now()+ datetime.timedelta(hours=-1)


SQL_QUERY = f"""

select * from TPABRM.BPART where P_CREATE_DATE>='2026/05/19 14:44'

"""

mykey=SQL_QUERY.split('from')[1].replace(' ','')[:2]
mykey1=SQL_QUERY.split('from')[1].split('\n')[0].split('.')[-1].split(' ')[0]
if mykey=='F7':
    conn=ibm_db_dbi.Connection(cnxncf)
elif mykey=='T7':
    conn=ibm_db_dbi.Connection(cnxntft)
elif mykey=='L7':
    conn=ibm_db_dbi.Connection(cnxnlcd)
elif mykey=='TP':
    conn=ibm_db_dbi.Connection(cnxntft)
else:
    print("請新增DB")
df = pd.read_sql(SQL_QUERY, conn)


myfilename=rf'T:\Software\WebSphere MQ 6.0_Client\Licenses\Windows\temp\{mydic[mykey]}_{mykey1}.xlsx'

with pd.ExcelWriter(myfilename, engine='openpyxl', mode='w') as writer:
    df.to_excel(writer, sheet_name=SQL_QUERY.split('from')[1].split('\n')[0].split('.')[-1].split(' ')[0],index=False)

reset_col(myfilename)