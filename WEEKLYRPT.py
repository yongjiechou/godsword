
import pandas as pd
import os,time,re,datetime,clipboard

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

def reset_col(filename):
	wb=load_workbook(filename)
	for sheet in wb.sheetnames:
		ws=wb[sheet]
		df=pd.read_excel(filename,sheet).fillna('-')
		df.loc[len(df)]=list(df.columns)						#把標題行附加到最後一行
		for col in df.columns:
			index=list(df.columns).index(col)					#列序號
			letter=get_column_letter(index+1)					#列字母
			collen=df[col].apply(lambda x:len(str(x).encode())).max()	#獲取這一列長度的最大值 當然也可以用min獲取最小值 mean獲取平均值
			if letter=='H':
				ws.column_dimensions[letter].width=collen*0.1+4
			else:
				ws.column_dimensions[letter].width=collen*1.1+4		#也就是列寬為最大長度*1.2 可以自己調整
##			print(letter,collen,collen*0.5+4,collen*1.2+4)
	wb.save(filename)
##os.add_dll_directory(r'D:\Portable Python-3.9.9 x64\App\Python\Lib\site-packages\clidriver\bin')
os.add_dll_directory(r'Z:\Software\WebSphere MQ 6.0_Client\Licenses\Windows\Portable Python-3.9.9 x64\App\Python\Lib\site-packages\clidriver\bin')

import ibm_db
import ibm_db_dbi

pd.set_option('display.max_rows', None)
pd.set_option('display.max_columns', None)
##pd.set_option('display.width', None) ##視情況調適欄位對應好看
pd.set_option('display.max_colwidth', None)


week=datetime.datetime.now().isocalendar()[1]-1
##week=10
the_week=f'W{str(week).zfill(2)}'
day_of_week=1
first_day = datetime.datetime.strptime(f"{datetime.datetime.now().year}-W{week}-{day_of_week}", "%G-W%V-%u")
last_day = datetime.datetime.strptime(f"{datetime.datetime.now().year}-W{week+1}-{day_of_week}", "%G-W%V-%u")

mydic={
'T7':'ARRAY7',
'F7':'CF7',
'L7':'LCD7',


}
cnxntft = ibm_db.connect("DATABASE=T7WPPT1;HOSTNAME=10.107.1.1;PORT=50101;PROTOCOL=TCPIP;UID=t7wbm1u2;PWD=t7insert;", "", "")
cnxncf = ibm_db.connect("DATABASE=F7WPPT1;HOSTNAME=10.107.1.2;PORT=50201;PROTOCOL=TCPIP;UID=f7wbm1u2;PWD=f7insert;", "", "")
cnxnlcd = ibm_db.connect("DATABASE=L7WPPT1;HOSTNAME=10.107.1.3;PORT=50301;PROTOCOL=TCPIP;UID=l7wbm1u2;PWD=l7insert;", "", "")

SQL_QUERY = f"""

select PRODUCT_ID,P_CREATE_DATE from T7WBM1D.BAPRDCT where P_CREATE_DATE>='{first_day.year}/{str(first_day.month).zfill(2)}/{str(first_day.day).zfill(2)} 00:00' and P_CREATE_DATE<='{last_day.year}/{str(last_day.month).zfill(2)}/{str(last_day.day).zfill(2)} 00:00'



"""
mykey=SQL_QUERY.split('from')[1].replace(' ','')[:2]
mykey1=SQL_QUERY.split('from')[1].split('\n')[0].split('.')[-1].split(' ')[0]
if mykey=='F7':
    conn=ibm_db_dbi.Connection(cnxncf)
elif mykey=='T7':
    conn=ibm_db_dbi.Connection(cnxntft)
elif mykey=='L7':
    conn=ibm_db_dbi.Connection(cnxnlcd)
elif mykey=='F5':
    conn=ibm_db_dbi.Connection(cnxncf5)
else:
    print("請新增DB")
df_DBT = pd.read_sql(SQL_QUERY, conn)

SQL_QUERY = f"""

select PFCD,P_CREATE_DATE from L7WBM1D.BCPRDCT where P_CREATE_DATE>='{first_day.year}/{str(first_day.month).zfill(2)}/{str(first_day.day).zfill(2)} 00:00' and P_CREATE_DATE<='{last_day.year}/{str(last_day.month).zfill(2)}/{str(last_day.day).zfill(2)} 00:00'


"""
mykey=SQL_QUERY.split('from')[1].replace(' ','')[:2]
mykey1=SQL_QUERY.split('from')[1].split('\n')[0].split('.')[-1].split(' ')[0]
if mykey=='F7':
    conn=ibm_db_dbi.Connection(cnxncf)
elif mykey=='T7':
    conn=ibm_db_dbi.Connection(cnxntft)
elif mykey=='L7':
    conn=ibm_db_dbi.Connection(cnxnlcd)
elif mykey=='F5':
    conn=ibm_db_dbi.Connection(cnxncf5)
else:
    print("請新增DB")
df_DBL = pd.read_sql(SQL_QUERY, conn)
df_DBL.columns=['PRODUCT_ID', 'P_CREATE_DATE']
df_DB = pd.concat([df_DBT, df_DBL], axis=0)
df_DB = df_DB.sort_values(by='P_CREATE_DATE')

myfile_a=r'Z:\Software\WebSphere MQ 6.0_Client\Licenses\Windows\download\週報.xlsx'
df_a = pd.read_excel(myfile_a,sheet_name=0)
df_b = pd.read_excel(myfile_a,sheet_name=1)
df_a=df_a.fillna("")
df_b=df_b.fillna("")
##print(df_a.head())

myfile_b=rf'Z:\Software\WebSphere MQ 6.0_Client\Licenses\Windows\download\{the_week}週報.xlsx'

myfilename=r'Z:\Software\WebSphere MQ 6.0_Client\Licenses\Windows\work\2025周雍傑週報.xlsx'
df = pd.read_excel(myfilename, sheet_name="7.簽核文件")
df=df.fillna("")

mydata=df[df['Week']==week]
datacount=len(mydata)
df_a.iloc[0,3]=datacount
df_a.iloc[0,4]=f'新產品{len(df_DB)}隻'
df_a.iloc[0,0]=the_week
mydatareturn=mydata[mydata['週會日期']!='']
datareturncount=len(mydatareturn)
df_a.iloc[1,3]=datareturncount
lastwip=mydata['ENG'].iloc[-1].split(':')[-1].replace('張','')
##df_a.iloc[2,3]=lastwip
df_a.iloc[1, 4] = ", ".join(map(str, lastwip))
unique_list = mydata['週會日期'].dropna().unique().tolist()
del unique_list[0]
##df_a.at[df_a.index[1], df_a.columns[4]] = unique_list
df_a.iloc[1, 4] = ", ".join(map(str, unique_list))
df_b=df_DB
with pd.ExcelWriter(myfile_b, engine='xlsxwriter') as writer:
    df_a.to_excel(writer, index=False, sheet_name='總表')
    df_b.to_excel(writer, index=False, sheet_name='新產品')

    workbook  = writer.book
    worksheet = writer.sheets['總表']
    worksheetb = writer.sheets['新產品']

    # --- 定義樣式 ---
    # 標題樣式：綠色背景 + 粗體 + 框線 + 置中
    header_format = workbook.add_format({
        'bold': True,
        'bg_color': '#C6E0B4',
        'border': 1,
        'align': 'center',
        'valign': 'vcenter'
    })

    # 一般單元格樣式：框線 + 置中 + 自動換行
    cell_format = workbook.add_format({
        'border': 1,
        'align': 'center',
        'valign': 'vcenter',
        'text_wrap': True
    })

    # --- 套用格式 ---
    # 重寫標題列
    for col_num, value in enumerate(df_a.columns.values):
        worksheet.write(0, col_num, value, header_format)
    for col_num, value in enumerate(df_b.columns.values):
        worksheetb.write(0, col_num, value, header_format)
    # 設置欄寬（讓備註欄寬一點）
    worksheet.set_column('A:D', 10, cell_format)
    worksheet.set_column('E:E', 40, cell_format) # 備註欄
    worksheet.set_column('F:F', 15, cell_format)
    worksheetb.set_column('A:C', 20, cell_format)

    # --- 關鍵：合併儲存格 ---
    # 語法：merge_range(開始列, 開始欄, 結束列, 結束欄, 內容, 樣式)
    # 注意：Excel 索引從 0 開始，這裡合併第 2 到 4 列 (索引 1~3)
    worksheet.merge_range(1, 0, 3, 0, f'{the_week}', cell_format)  # 合併週別
    worksheet.merge_range(1, 1, 3, 1, 'FAB7', cell_format) # 合併廠別
    worksheet.merge_range(1, 5, 3, 5, '周雍傑', cell_format) # 合併負責人
